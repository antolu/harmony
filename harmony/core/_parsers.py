"""Document parsers for extracting text from various file formats."""

from __future__ import annotations

import csv
import io
import typing
from pathlib import Path
from xml.dom import minidom

import chardet
import docx
from odf import opendocument  # type: ignore[import-untyped]  # odfpy has no stubs
from odf import text as odf_text
from openpyxl import (  # type: ignore[import-untyped]  # openpyxl has no stubs
    load_workbook,
)
from pypdf import PdfReader

if typing.TYPE_CHECKING:
    pass


class ParseError(Exception):
    """Base exception for document parsing errors."""


class UnsupportedDocumentError(ParseError):
    """Document format not supported."""


class CorruptDocumentError(ParseError):
    """Document is corrupted or malformed."""


class DocumentParser(typing.Protocol):
    """Protocol for document parsers."""

    def can_parse(self, content_type: str, extension: str) -> bool:
        """Check if parser can handle this document type."""
        ...

    def parse(self, filepath: Path) -> tuple[str, str]:
        """Parse document and return (title, content).

        Args:
            filepath: Path to document file

        Returns:
            Tuple of (title, content)

        Raises:
            ParseError: If parsing fails
        """
        ...


class PDFParser:
    """Parser for PDF documents using pypdf."""

    def can_parse(self, content_type: str, extension: str) -> bool:
        """Check if this is a PDF."""
        return content_type == "application/pdf" or extension == ".pdf"

    def parse(self, filepath: Path) -> tuple[str, str]:
        """Extract text from PDF."""
        try:
            return self._parse(filepath)
        except Exception as e:
            msg = f"Failed to parse PDF: {e}"
            raise CorruptDocumentError(msg) from e

    def _parse(self, filepath: Path) -> tuple[str, str]:
        reader = PdfReader(filepath)
        title = reader.metadata.title if reader.metadata else ""
        if not title:
            title = filepath.stem

        content_parts = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                content_parts.append(text)

        content = "\n\n".join(content_parts)
        return title, content


class DocxParser:
    """Parser for Microsoft Word DOCX documents."""

    def can_parse(self, content_type: str, extension: str) -> bool:
        """Check if this is a DOCX."""
        return (
            content_type
            == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            or extension == ".docx"
        )

    def parse(self, filepath: Path) -> tuple[str, str]:
        """Extract text from DOCX."""
        try:
            return self._parse(filepath)
        except Exception as e:
            msg = f"Failed to parse DOCX: {e}"
            raise CorruptDocumentError(msg) from e

    def _parse(self, filepath: Path) -> tuple[str, str]:
        doc = docx.Document(str(filepath))
        title = doc.core_properties.title or filepath.stem
        paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]
        content = "\n\n".join(paragraphs)
        return title, content


class XlsxParser:
    """Parser for Microsoft Excel XLSX spreadsheets."""

    def can_parse(self, content_type: str, extension: str) -> bool:
        """Check if this is an XLSX."""
        return (
            content_type
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            or extension == ".xlsx"
        )

    def parse(self, filepath: Path) -> tuple[str, str]:
        """Extract text from XLSX."""
        try:
            return self._parse(filepath)
        except Exception as e:
            msg = f"Failed to parse XLSX: {e}"
            raise CorruptDocumentError(msg) from e

    def _parse(self, filepath: Path) -> tuple[str, str]:
        workbook = load_workbook(filepath, read_only=True, data_only=True)
        title = filepath.stem
        content_parts = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            content_parts.append(f"Sheet: {sheet_name}")
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join(str(cell) for cell in row if cell is not None)
                if row_text.strip():
                    content_parts.append(row_text)
        content = "\n".join(content_parts)
        return title, content


class OdtParser:
    """Parser for OpenDocument Text (ODT) documents."""

    def can_parse(self, content_type: str, extension: str) -> bool:
        """Check if this is an ODT."""
        return (
            content_type == "application/vnd.oasis.opendocument.text"
            or extension == ".odt"
        )

    def parse(self, filepath: Path) -> tuple[str, str]:
        """Extract text from ODT."""
        try:
            return self._parse(filepath)
        except Exception as e:
            msg = f"Failed to parse ODT: {e}"
            raise CorruptDocumentError(msg) from e

    def _parse(self, filepath: Path) -> tuple[str, str]:
        doc = opendocument.load(str(filepath))
        title = ""
        meta = doc.meta
        if meta:
            title_elements = meta.getElementsByType(odf_text.Title)
            if title_elements:
                title = str(title_elements[0])
        if not title:
            title = filepath.stem
        paragraphs = []
        for paragraph in doc.getElementsByType(odf_text.P):
            text_content = self._extract_text_from_element(paragraph)
            if text_content.strip():
                paragraphs.append(text_content)
        content = "\n\n".join(paragraphs)
        return title, content

    def _extract_text_from_element(self, element: minidom.Node) -> str:
        """Recursively extract text from ODF element."""
        text_parts = []
        for child in element.childNodes:
            if hasattr(child, "data"):
                text_parts.append(child.data)
            elif hasattr(child, "childNodes"):
                text_parts.append(self._extract_text_from_element(child))
        return "".join(text_parts)


class TextParser:
    """Parser for plain text files with encoding detection."""

    def can_parse(self, content_type: str, extension: str) -> bool:
        """Check if this is a text file."""
        return content_type.startswith("text/") or extension in {
            ".txt",
            ".text",
            ".rtf",
        }

    def parse(self, filepath: Path) -> tuple[str, str]:
        """Extract text from text file with encoding detection."""
        try:
            return self._parse(filepath)
        except Exception as e:
            msg = f"Failed to parse text file: {e}"
            raise CorruptDocumentError(msg) from e

    def _parse(self, filepath: Path) -> tuple[str, str]:
        raw_data = filepath.read_bytes()
        detected = chardet.detect(raw_data)
        encoding = detected["encoding"] or "utf-8"
        content = raw_data.decode(encoding, errors="replace")
        title = filepath.stem
        return title, content


class MarkdownParser:
    """Parser for Markdown files."""

    def can_parse(self, content_type: str, extension: str) -> bool:
        """Check if this is a Markdown file."""
        return extension in {".md", ".markdown", ".mdown", ".mkd"} or content_type in {
            "text/markdown",
            "text/x-markdown",
        }

    def parse(self, filepath: Path) -> tuple[str, str]:
        """Extract text from Markdown file and extract title from first H1."""
        try:
            return self._parse(filepath)
        except Exception as e:
            msg = f"Failed to parse Markdown file: {e}"
            raise CorruptDocumentError(msg) from e

    def _parse(self, filepath: Path) -> tuple[str, str]:
        raw_data = filepath.read_bytes()
        detected = chardet.detect(raw_data)
        encoding = detected["encoding"] or "utf-8"
        content = raw_data.decode(encoding, errors="replace")
        title = filepath.stem
        lines = content.split("\n")
        for i, line in enumerate(lines):
            if line.startswith("# "):
                title = line[2:].strip()
                break
            if i > 0 and line.strip() and all(c == "=" for c in line.strip()):
                title = lines[i - 1].strip()
                break
        return title, content


class CsvParser:
    """Parser for CSV files."""

    def can_parse(self, content_type: str, extension: str) -> bool:
        """Check if this is a CSV."""
        return content_type == "text/csv" or extension == ".csv"

    def parse(self, filepath: Path) -> tuple[str, str]:
        """Extract text from CSV."""
        try:
            return self._parse(filepath)
        except Exception as e:
            msg = f"Failed to parse CSV: {e}"
            raise CorruptDocumentError(msg) from e

    def _parse(self, filepath: Path) -> tuple[str, str]:
        raw_data = filepath.read_bytes()
        detected = chardet.detect(raw_data)
        encoding = detected["encoding"] or "utf-8"
        content_text = raw_data.decode(encoding, errors="replace")
        reader = csv.reader(io.StringIO(content_text))
        rows = []
        for row in reader:
            row_text = "\t".join(cell for cell in row if cell)
            if row_text.strip():
                rows.append(row_text)
        content = "\n".join(rows)
        title = filepath.stem
        return title, content


class ParserRegistry:
    """Registry for document parsers."""

    def __init__(self) -> None:
        self.parsers: list[DocumentParser] = []
        self._register_default_parsers()

    def _register_default_parsers(self) -> None:
        """Register all default parsers."""
        self.register(PDFParser())
        self.register(DocxParser())
        self.register(XlsxParser())
        self.register(OdtParser())
        self.register(CsvParser())
        self.register(MarkdownParser())
        self.register(TextParser())

    def register(self, parser: DocumentParser) -> None:
        """Register a parser."""
        self.parsers.append(parser)

    def get_parser(self, content_type: str, extension: str) -> DocumentParser | None:
        """Get parser for content type and extension."""
        for parser in self.parsers:
            if parser.can_parse(content_type, extension):
                return parser
        return None


default_registry = ParserRegistry()
